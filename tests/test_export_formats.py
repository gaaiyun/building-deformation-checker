"""src.tools.export_formats 单元测试

构造最小化的 MonitoringReport，验证 DOCX/HTML 导出能产生合法二进制/标记。
不依赖任何外部 LLM/OCR 服务，纯本地确定性测试。
"""

from __future__ import annotations

import io
import sys
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.models.data_models import (
    CheckIssue,
    MeasurementPoint,
    MonitoringCategory,
    MonitoringReport,
    MonitoringTable,
)
from src.tools.export_formats import generate_docx, generate_html, generate_intermediate_xlsx


def _make_minimal_report(project_name: str = "测试项目") -> MonitoringReport:
    """构造一个最小化但字段完整的 MonitoringReport"""
    table = MonitoringTable(
        monitoring_item="水平位移",
        category=MonitoringCategory.HORIZONTAL_DISP,
        monitor_date="2026-05-17",
        point_count=1,
        points=[
            MeasurementPoint(
                point_id="P1",
                initial_value=0.0,
                previous_value=0.5,
                current_value=1.0,
                current_change=0.5,
                cumulative_change=1.0,
                change_rate=0.1,
                safety_status="正常",
            )
        ],
    )
    return MonitoringReport(
        project_name=project_name,
        monitoring_company="测试单位",
        report_number="ZB-2026-001",
        monitoring_date="2026-05-17",
        tables=[table],
        conclusion="本期监测数据正常。",
        extraction_diagnostics={
            "method": "text_layer",
            "selected_profile": "default",
            "raw_chars": 12345,
            "clean_chars": 10000,
            "compression_ratio": 0.81,
            "abnormal_table_count": 0,
        },
    )


def _make_issue(severity: str = "error", msg: str = "测试问题") -> CheckIssue:
    return CheckIssue(
        severity=severity,
        table_name="水平位移",
        point_id="P1",
        field_name="cumulative_change",
        expected_value="1.0",
        actual_value="2.0",
        message=msg,
    )


class TestGenerateDocx(unittest.TestCase):
    """generate_docx 应产生合法的 .docx (ZIP) 二进制内容"""

    def test_returns_bytes(self):
        report = _make_minimal_report()
        out = generate_docx("# 报告\n\n正文", report, errors=[], warnings=[])
        self.assertIsInstance(out, bytes)
        self.assertGreater(len(out), 0)

    def test_starts_with_zip_signature(self):
        """.docx 本质是 ZIP 容器，应以 PK 头开头"""
        report = _make_minimal_report()
        out = generate_docx("# 报告", report, errors=[], warnings=[])
        # ZIP 文件签名: 0x50 0x4B 0x03 0x04
        self.assertEqual(out[:2], b"PK")

    def test_is_valid_zip_archive(self):
        report = _make_minimal_report()
        out = generate_docx("# 报告", report, errors=[], warnings=[])
        buf = io.BytesIO(out)
        self.assertTrue(zipfile.is_zipfile(buf))
        with zipfile.ZipFile(buf) as zf:
            names = zf.namelist()
            # docx 标准结构应包含
            self.assertIn("word/document.xml", names)
            self.assertIn("[Content_Types].xml", names)

    def test_empty_errors_and_warnings_dont_crash(self):
        report = _make_minimal_report()
        # 既无错误也无警告
        out = generate_docx("# 报告", report, errors=[], warnings=[])
        self.assertIsInstance(out, bytes)
        self.assertGreater(len(out), 1000)  # 至少有些字节

    def test_with_errors_and_warnings(self):
        report = _make_minimal_report()
        errors = [_make_issue("error", "错误1"), _make_issue("error", "错误2")]
        warnings = [_make_issue("warning", "警告1")]
        out = generate_docx("# 报告", report, errors=errors, warnings=warnings)
        self.assertIsInstance(out, bytes)
        # 解开 docx 看 document.xml 是否包含错误信息
        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            doc_xml = zf.read("word/document.xml").decode("utf-8")
        self.assertIn("错误1", doc_xml)
        self.assertIn("警告1", doc_xml)

    def test_microsoft_yahei_font_in_xml(self):
        """生成的 docx 应该把字体设为 Microsoft YaHei (中文字符防回退方块)"""
        report = _make_minimal_report()
        out = generate_docx("# 报告", report, errors=[], warnings=[])
        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            doc_xml = zf.read("word/document.xml").decode("utf-8")
            styles_xml = ""
            try:
                styles_xml = zf.read("word/styles.xml").decode("utf-8")
            except KeyError:
                pass
        combined = doc_xml + styles_xml
        self.assertIn("Microsoft YaHei", combined)

    def test_project_name_appears_in_overview(self):
        report = _make_minimal_report(project_name="某某基坑监测")
        out = generate_docx("", report, errors=[], warnings=[])
        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            doc_xml = zf.read("word/document.xml").decode("utf-8")
        self.assertIn("某某基坑监测", doc_xml)


class TestGenerateHtml(unittest.TestCase):
    """generate_html 应产生合法 HTML5 字符串"""

    def test_returns_string(self):
        html = generate_html("# 标题\n\n正文段落", "测试项目")
        self.assertIsInstance(html, str)
        self.assertGreater(len(html), 0)

    def test_has_html5_doctype(self):
        html = generate_html("# 标题", "测试项目")
        self.assertTrue(html.lstrip().startswith("<!DOCTYPE html>"))

    def test_has_html_root_element(self):
        html = generate_html("# 标题", "测试项目")
        self.assertIn("<html", html)
        self.assertIn("</html>", html)

    def test_has_zh_cn_lang(self):
        html = generate_html("# 标题", "测试项目")
        self.assertIn('lang="zh-CN"', html)

    def test_contains_project_name_in_title(self):
        html = generate_html("正文", "我的工程项目XYZ")
        self.assertIn("我的工程项目XYZ", html)
        # 应在 <title> 中
        self.assertIn("<title>", html)

    def test_renders_markdown_to_html(self):
        md = "# 一级标题\n\n## 二级标题\n\n- 列表项"
        html = generate_html(md, "项目")
        # markdown 标题应转为 <h1>/<h2>
        self.assertIn("<h1>", html)
        self.assertIn("一级标题", html)

    def test_markdown_table_renders(self):
        md = "| A | B |\n|---|---|\n| 1 | 2 |\n"
        html = generate_html(md, "项目")
        self.assertIn("<table>", html)
        self.assertIn("<th>", html)

    def test_empty_markdown_does_not_crash(self):
        html = generate_html("", "项目X")
        self.assertIsInstance(html, str)
        self.assertIn("<!DOCTYPE html>", html)
        self.assertIn("项目X", html)

    def test_microsoft_yahei_font_in_css(self):
        html = generate_html("# 标题", "项目")
        self.assertIn("Microsoft YaHei", html)

    def test_print_media_styles_included(self):
        """HTML 报告应支持打印为 PDF，包含 @media print"""
        html = generate_html("# 标题", "项目")
        self.assertIn("@media print", html)

    def test_has_charset_utf8(self):
        html = generate_html("# 中文标题", "项目")
        self.assertIn('charset="UTF-8"', html.upper().replace('CHARSET="UTF-8"', 'charset="UTF-8"'))


class TestGenerateIntermediateXlsx(unittest.TestCase):
    """generate_intermediate_xlsx 应产生可审查的 Excel 中间层"""

    def test_returns_valid_xlsx(self):
        report = _make_minimal_report()
        out = generate_intermediate_xlsx(report)
        self.assertIsInstance(out, bytes)
        self.assertEqual(out[:2], b"PK")
        wb = load_workbook(io.BytesIO(out), read_only=True)
        self.assertIn("00_报告概览", wb.sheetnames)
        self.assertIn("01_表格清单", wb.sheetnames)
        self.assertIn("02_标准化测点", wb.sheetnames)
        self.assertIn("05_问题清单", wb.sheetnames)

    def test_standardized_point_sheet_contains_measurement_fields(self):
        report = _make_minimal_report(project_name="Excel中间层项目")
        out = generate_intermediate_xlsx(report)
        wb = load_workbook(io.BytesIO(out), read_only=True, data_only=True)
        ws = wb["02_标准化测点"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][:5], ("表序号", "表名", "类别", "日期", "测点"))
        self.assertIn("P1", rows[1])
        self.assertIn(1.0, rows[1])

    def test_issue_sheet_contains_all_issue_sources(self):
        report = _make_minimal_report()
        out = generate_intermediate_xlsx(
            report,
            calc_issues=[_make_issue("error", "计算错误")],
            stats_issues=[_make_issue("warning", "统计警告")],
            logic_issues=[_make_issue("info", "逻辑提示")],
        )
        wb = load_workbook(io.BytesIO(out), read_only=True, data_only=True)
        ws = wb["05_问题清单"]
        body = ["|".join(str(v or "") for v in row) for row in ws.iter_rows(values_only=True)]
        joined = "\n".join(body)
        self.assertIn("计算核验", joined)
        self.assertIn("统计核验", joined)
        self.assertIn("逻辑检查", joined)
        self.assertIn("计算错误", joined)

    def test_text_cells_escape_excel_formula_prefixes(self):
        report = _make_minimal_report(project_name="=危险项目")
        report.tables[0].points[0].point_id = "=SUM(1,1)"
        issue = _make_issue("error", "@危险说明")
        out = generate_intermediate_xlsx(report, calc_issues=[issue])

        wb = load_workbook(io.BytesIO(out), read_only=True, data_only=True)
        overview = wb["00_报告概览"]
        point_ws = wb["02_标准化测点"]
        issue_ws = wb["05_问题清单"]

        self.assertEqual(overview["B2"].value, "'=危险项目")
        self.assertEqual(point_ws["E2"].value, "'=SUM(1,1)")
        self.assertIn("'@危险说明", issue_ws["H2"].value)


if __name__ == "__main__":
    unittest.main()
