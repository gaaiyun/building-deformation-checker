"""逐 cell 比较 错误版 vs 正确版，输出 ground truth diff。

输出：
  baseline/diffs/<company>_diff.md  - markdown diff 清单
  baseline/diffs/<company>_diff.json - 结构化 diff（程序可读）
"""

from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path
from typing import Any

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DIFFS_DIR = ROOT / "baseline" / "diffs"
DIFFS_DIR.mkdir(parents=True, exist_ok=True)

PAIRS = [
    ("质安", "质安模板-错误版.xlsx", "质安模板-正确版.xlsx"),
    ("深工勘", "深工勘模板-错误版.xlsx", "深工勘模板-正确版.xlsx"),
    ("展誉", "展誉模板-错误版.xlsx", "展誉模板-正确版.xlsx"),
]


def cell_diffs(err_path: Path, cor_path: Path) -> dict[str, list[dict]]:
    """每 sheet 找差异 cell。data_only=True 拿到公式计算值。

    每条 diff 还会带上行内的"测点编号"（A 列）与列上方的"字段名"（header
    row），方便人和工具都能理解 "F11" 实际是 'WY236 的变化速率列'。
    """
    err_wb = load_workbook(str(err_path), data_only=True)
    cor_wb = load_workbook(str(cor_path), data_only=True)

    out: dict[str, list[dict]] = {}
    for sheet_name in err_wb.sheetnames:
        if sheet_name not in cor_wb.sheetnames:
            continue
        ews = err_wb[sheet_name]
        cws = cor_wb[sheet_name]

        # 找 header 行（含"测点编号"字样的行）
        header_row_idx = _find_header_row(cws)
        col_headers = _extract_col_headers(cws, header_row_idx) if header_row_idx else {}

        diffs = []
        max_r = max(ews.max_row, cws.max_row)
        max_c = max(ews.max_column, cws.max_column)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                e_val = ews.cell(r, c).value
                c_val = cws.cell(r, c).value
                if _normalize(e_val) != _normalize(c_val):
                    coord = f"{get_column_letter(c)}{r}"
                    # 行上下文：A 列 cell 值（通常是测点编号）
                    point_id = _stringify(cws.cell(r, 1).value)
                    # 列上下文：header 行对应的字段名
                    field_name = col_headers.get(c, "")
                    diffs.append({
                        "sheet": sheet_name,
                        "cell": coord,
                        "row": r,
                        "col": c,
                        "point_id": point_id if point_id != "(空)" else "",
                        "field_name": field_name,
                        "error_value": _stringify(e_val),
                        "correct_value": _stringify(c_val),
                        "diff_type": _classify(e_val, c_val),
                    })
        if diffs:
            out[sheet_name] = diffs

    err_wb.close()
    cor_wb.close()
    return out


def _find_header_row(ws) -> int:
    """启发式找 header 行：包含 '测点编号' 字样的行"""
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, min(ws.max_column, 5) + 1):
            v = ws.cell(r, c).value
            if v and "测点编号" in str(v):
                return r
    return 0


def _extract_col_headers(ws, header_row: int) -> dict[int, str]:
    """从 header 行（及紧邻下方 1-2 行）抽取列名映射"""
    headers = {}
    for c in range(1, ws.max_column + 1):
        parts = []
        for r in range(header_row, min(header_row + 3, ws.max_row + 1)):
            v = ws.cell(r, c).value
            if v and str(v).strip():
                parts.append(str(v).strip())
        if parts:
            headers[c] = " / ".join(parts)
    return headers


def _normalize(v: Any) -> Any:
    """归一以避免 false positive 差异：
    - None / 空串 → 统一为 ""
    - 数值（含可解析为数字的字符串）→ 四舍五入到 6 位小数
    - 字符串去首尾空白
    """
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        try:
            return round(float(v), 6)
        except (ValueError, OverflowError):
            return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return ""
        # 尝试当数字解析（XLSX 偶尔把数字以字符串形式存）
        try:
            return round(float(s), 6)
        except ValueError:
            return s
    return v


def _stringify(v: Any) -> str:
    if v is None:
        return "(空)"
    if isinstance(v, float):
        # 保留至多 6 位小数，去尾随 0
        s = f"{v:.6f}".rstrip("0").rstrip(".")
        return s or "0"
    return str(v)


def _classify(e: Any, c: Any) -> str:
    """启发式分类差异类型，归一化后比较"""
    en = _normalize(e)
    cn = _normalize(c)

    if cn == "" or cn is None:
        return "新增（错误版多了内容）"
    if en == "" or en is None:
        return "缺失（错误版漏了内容）"
    if isinstance(en, (int, float)) and isinstance(cn, (int, float)):
        delta = float(en) - float(cn)
        rel = abs(delta) / max(abs(float(cn)), 1e-9)
        return f"数值改动（差 {delta:+.4f}，相对 {rel:.1%}）"
    if isinstance(en, str) and isinstance(cn, str):
        return "文本改动"
    return "类型不一致"


def write_md(company: str, diffs_by_sheet: dict[str, list[dict]]) -> Path:
    """生成 markdown diff 报告"""
    total = sum(len(v) for v in diffs_by_sheet.values())
    lines = [
        f"# {company} 模板 · 错误版 vs 正确版 · 逐 cell Diff",
        "",
        f"**总差异数**: {total} 处",
        f"**涉及 sheet**: {len(diffs_by_sheet)} 个",
        "",
        "本文件由 `baseline/xlsx_diff.py` 自动生成，作为工具核查的 **ground truth**：",
        "理想情况下，v2 工具跑错误版应该恰好发现下面这些问题；跑正确版应该 0 错误。",
        "",
    ]

    for sheet, diffs in diffs_by_sheet.items():
        lines.append(f"## Sheet: `{sheet}` — {len(diffs)} 处差异")
        lines.append("")
        lines.append("| # | 单元格 | 测点编号 | 字段 | 错误版 | 正确版 | 差异类型 |")
        lines.append("|---|--------|----------|------|--------|--------|----------|")
        for i, d in enumerate(diffs, 1):
            err_val = d["error_value"].replace("|", "\\|")[:30]
            cor_val = d["correct_value"].replace("|", "\\|")[:30]
            pid = d.get("point_id", "").replace("|", "\\|")[:15]
            field = d.get("field_name", "").replace("|", "\\|")[:25]
            lines.append(
                f"| {i} | `{d['cell']}` | {pid} | {field} | {err_val} | {cor_val} | {d['diff_type']} |"
            )
        lines.append("")

    out = DIFFS_DIR / f"{company}_diff.md"
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def write_json(company: str, diffs_by_sheet: dict[str, list[dict]]) -> Path:
    out = DIFFS_DIR / f"{company}_diff.json"
    out.write_text(
        json.dumps(diffs_by_sheet, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out


for company, err_name, cor_name in PAIRS:
    err = ROOT / err_name
    cor = ROOT / cor_name
    print(f"\n{'=' * 70}")
    print(f"  {company}")
    print("=" * 70)
    if not err.exists() or not cor.exists():
        print(f"  ❌ 缺文件")
        continue

    diffs = cell_diffs(err, cor)
    total = sum(len(v) for v in diffs.values())
    print(f"  发现 {total} 处差异，涉及 {len(diffs)} 个 sheet：")
    for s, ds in diffs.items():
        print(f"    {s:30}  {len(ds):3} 处")

    md = write_md(company, diffs)
    js = write_json(company, diffs)
    print(f"  ✅ 输出: {md.name} + {js.name}")
