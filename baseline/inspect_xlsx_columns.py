"""检查深工勘/展誉 XLSX 各 sheet 的列布局，理解 GT 错误所在列的含义"""

from __future__ import annotations

import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]

# 我们要查的：根据 diff 中的 cell 坐标，倒推这个 cell 的列含义
QUERIES = [
    # (file, sheet, cells of interest)
    ("深工勘模板-正确版.xlsx", "支护结构顶部水平位移", ["E", "J", "T"], [4, 5, 6, 7, 8, 25, 45]),
    ("深工勘模板-正确版.xlsx", "周边地面沉降", ["M", "G", "O", "C", "E"], [4, 5, 10, 20, 21, 46]),
    ("深工勘模板-正确版.xlsx", "支护桩支护结构深层水平位移", ["F", "I", "J", "C"], [4, 5, 15, 16, 20, 21, 71, 80]),
    ("展誉模板-正确版.xlsx", "基坑顶水平位移", ["F", "C", "I"], [4, 5, 11, 12, 13, 16, 64]),
    ("展誉模板-正确版.xlsx", "立柱沉降", ["F", "H", "B", "C", "I"], [4, 5, 13, 14, 58]),
    ("展誉模板-正确版.xlsx", "地下水位", ["F", "I"], [4, 5, 8, 10, 56]),
    ("展誉模板-正确版.xlsx", "建筑物倾斜X", ["C", "F", "I"], [4, 5, 8, 9]),
]


for fname, sheet_name, cols, rows in QUERIES:
    print(f"\n{'=' * 70}")
    print(f"  {fname} :: {sheet_name}")
    print("=" * 70)
    path = ROOT / fname
    if not path.exists():
        print(f"  ❌ 缺文件")
        continue
    wb = load_workbook(str(path), data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  ❌ 无 sheet")
        continue
    ws = wb[sheet_name]

    # 打印表头（前 6 行）
    print("\n[表头 1-6 行]")
    for r in range(1, 7):
        row_data = []
        for c in range(1, min(ws.max_column + 1, 26)):
            v = ws.cell(r, c).value
            cell_id = f"{get_column_letter(c)}{r}"
            if v is not None and str(v).strip():
                row_data.append(f"{cell_id}={str(v)[:15]}")
        print(f"  R{r}: {' | '.join(row_data) if row_data else '(空)'}")

    print("\n[感兴趣的 cells 周围内容]")
    for r in rows:
        if r > ws.max_row:
            continue
        print(f"  R{r}:")
        # 打印第 r 行所有列（前 25 列）的值
        for c in range(1, min(ws.max_column + 1, 26)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                marker = " ★" if get_column_letter(c) in cols else ""
                print(f"    {get_column_letter(c)}{r} = {str(v)[:20]}{marker}")

    wb.close()
