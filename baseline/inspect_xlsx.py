"""探查 3 对 XLSX 模板的结构：sheet 数、每 sheet 的尺寸与首行/首列"""

from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FILES = [
    "质安模板-错误版.xlsx",
    "质安模板-正确版.xlsx",
    "深工勘模板-错误版.xlsx",
    "深工勘模板-正确版.xlsx",
    "展誉模板-错误版.xlsx",
    "展誉模板-正确版.xlsx",
]


def inspect(path: Path) -> None:
    print(f"\n{'=' * 70}")
    print(f"📄 {path.name}  ({path.stat().st_size / 1024:.1f} KB)")
    print("=" * 70)
    wb = load_workbook(str(path), read_only=True, data_only=False)
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"\n  [{idx}] Sheet: {sheet_name!r}  尺寸: {ws.max_row} 行 × {ws.max_column} 列")

        # 抓首 3 行非空 cell 内容
        sample_rows = []
        row_count = 0
        for row in ws.iter_rows(values_only=True, max_row=6):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            # 只取前 8 列 + 限制 cell 长度
            preview = [
                (str(c)[:25] + "..." if c and len(str(c)) > 25 else (c if c is not None else ""))
                for c in row[:8]
            ]
            sample_rows.append(preview)
            row_count += 1
            if row_count >= 5:
                break

        for i, row in enumerate(sample_rows, 1):
            print(f"      第 {i} 行 (前 8 列): {row}")

    wb.close()


for fname in FILES:
    path = ROOT / fname
    if not path.exists():
        print(f"❌ 未找到: {fname}")
        continue
    try:
        inspect(path)
    except Exception as e:
        print(f"❌ 读取失败 {fname}: {e}")
