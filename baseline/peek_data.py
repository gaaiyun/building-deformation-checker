"""快速浏览正确版 XLSX 的数据 sheet，理解多日期块结构"""

from __future__ import annotations

import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb = load_workbook(str(ROOT / "质安模板-正确版.xlsx"), data_only=True)
ws = wb["支护结构水平位移"]
print(f"Sheet: 支护结构水平位移, 尺寸: {ws.max_row} × {ws.max_column}")
print()

# 打印所有行内容（关键看怎么分块）
for r in range(1, min(ws.max_row + 1, 80)):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    non_empty = sum(1 for v in row if v is not None and str(v).strip())
    if non_empty == 0:
        continue
    cells_str = " | ".join(
        f"{(str(v)[:18] if v else ''):>18}" for v in row
    )
    print(f"  R{r:>3}  ne={non_empty}  | {cells_str}")
