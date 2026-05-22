"""探查 XLSX 结构 v2：非 read_only 模式，更准确"""

from __future__ import annotations

import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PAIRS = [
    ("质安模板-错误版.xlsx", "质安模板-正确版.xlsx"),
    ("深工勘模板-错误版.xlsx", "深工勘模板-正确版.xlsx"),
    ("展誉模板-错误版.xlsx", "展誉模板-正确版.xlsx"),
]


def summarize(path: Path) -> dict:
    """返回结构摘要：{sheet_name: {dims, sample_rows}}"""
    wb = load_workbook(str(path), data_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        # 计算真实尺寸（找最后非空 cell）
        max_r, max_c = 0, 0
        sample_rows = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            non_empty = [(i, c) for i, c in enumerate(row) if c is not None and str(c).strip()]
            if non_empty:
                max_r = max(max_r, row_idx)
                max_c = max(max_c, non_empty[-1][0] + 1)
                if len(sample_rows) < 5:
                    sample_rows.append(
                        [(str(c)[:30] if c is not None else "") for c in row[:10]]
                    )
        out[name] = {
            "rows": max_r,
            "cols": max_c,
            "sample": sample_rows,
        }
    wb.close()
    return out


for err_name, cor_name in PAIRS:
    err_path = ROOT / err_name
    cor_path = ROOT / cor_name
    print(f"\n{'#' * 75}")
    print(f"# 对比对: {err_name}  vs  {cor_name}")
    print("#" * 75)

    print("\n[读取 ERROR 版]")
    err = summarize(err_path)
    print("\n[读取 CORRECT 版]")
    cor = summarize(cor_path)

    # 对齐：列出双方的 sheet
    all_sheets = list(dict.fromkeys(list(err.keys()) + list(cor.keys())))
    print(f"\n  错误版 sheet: {list(err.keys())}")
    print(f"  正确版 sheet: {list(cor.keys())}")
    print(f"  合并清单 ({len(all_sheets)} 个):")
    for name in all_sheets:
        e_dims = (err[name]["rows"], err[name]["cols"]) if name in err else "缺失"
        c_dims = (cor[name]["rows"], cor[name]["cols"]) if name in cor else "缺失"
        diff = "✓ 同尺寸" if e_dims == c_dims else f"⚠ 错={e_dims} 正={c_dims}"
        print(f"    {name:30}  错误版={e_dims}  正确版={c_dims}  {diff}")
